"""
Migracion xlsx -> SQLite  (modelo de cruces ferroviarios L2)
============================================================
Lee los dos libros Excel originales y construye las tres bases de datos:
    data/infraestructura.db   data/demanda.db   data/escenarios.db

Uso:
    python scripts/migrar_xlsx.py

Requiere que los .xlsx esten en la carpeta  fuentes/  (ver README).
Es idempotente: recrea las bases desde cero en cada corrida.
"""
from __future__ import annotations
import datetime
import sqlite3
import unicodedata
from pathlib import Path

import openpyxl

RAIZ      = Path(__file__).resolve().parent.parent
DIR_FUENTES = RAIZ / 'fuentes'
DIR_DATA  = RAIZ / 'data'
DIR_SCHEMA = DIR_DATA / 'schema'

XLSX_BASE   = 'Analisis_Cruces_L2_NOREPROG_PREVACIADO_PLAN_DINAMICO_BASE_REAL.xlsx'
XLSX_REPROG = 'Analisis_Cruces_L2_Reprog_Mar9_PREVACIADO_N2_PLAN_DINAMICO_REHECHO.xlsx'

# cruce -> (fila HCALL-IN, fila HCALL-OUT) en la hoja HCALL

# Cruces definidos por el proyecto para operar con reconfiguracion.
CRUCES_RECONFIGURACION = {
    'Los Claveles', 'Diagonal Bio Bio', 'Michaihue', 'Masisa',
    'Lomas Coloradas', 'Portal San Pedro', 'Conavicop', 'Conavicoop',
    'Escuadron 2',
}
DESC_RECONFIG = 'Base + HCALL + prevaciado + reconfiguracion'
DESC_NOREPROG = 'Base + HCALL + prevaciado + noreprog/base'

HCALL_ROWS = {
    'Conavicop': (15, 41), 'Portal San Pedro': (16, 42),
    'Lomas Coloradas': (17, 43), 'Costa Verde': (19, 45),
    'Michaihue': (20, 46), 'Diagonal Bio Bio': (22, 48),
    'Los Claveles': (24, 50),
}


# --------------------------------------------------------------------- #
#  utilidades de parsing
# --------------------------------------------------------------------- #
def to_seconds(v):
    """Convierte hora/tiempo de una celda a segundos enteros (o None)."""
    if v is None:
        return None
    if isinstance(v, (datetime.datetime, datetime.time)):
        return v.hour * 3600 + v.minute * 60 + v.second
    if isinstance(v, (int, float)):
        return int(round(v * 86400)) if 0 <= v < 2 else int(round(v))
    s = str(v).strip()
    if s in ('', '#DIV/0!', '#NAME?', '#VALUE!'):
        return None
    if ':' in s:
        p = [float(x) for x in s.split(':')]
        while len(p) < 3:
            p.append(0)
        return int(round(p[0] * 3600 + p[1] * 60 + p[2]))
    try:
        return int(round(float(s.replace(',', '.'))))
    except ValueError:
        return None


def to_float_cl(v):
    """Convierte numero con coma decimal chilena a float."""
    if v is None:
        return None
    if isinstance(v, (int, float)):
        return float(v)
    try:
        return float(str(v).strip().replace(',', '.'))
    except ValueError:
        return None


def norm(s):
    """Normaliza nombres para emparejar (sin acentos, minusculas)."""
    if s is None:
        return ''
    s = unicodedata.normalize('NFKD', str(s)).encode('ascii', 'ignore').decode()
    return ' '.join(s.lower().split())



def poblar_modelo_operacional(con: sqlite3.Connection) -> None:
    # Puebla la asignacion operacional por cruce.
    # Los cruces definidos por proyecto usan la version reprogramada; el resto
    # usa la version base/noreprog. Esto evita tratar la reconfiguracion como un
    # selector global de escenario.
    reconfig_norm = {norm(x) for x in CRUCES_RECONFIGURACION}
    for row in con.execute('SELECT cruce_id, nombre FROM cruces ORDER BY cruce_id').fetchall():
        cruce_id, nombre = row[0], row[1]
        usa = norm(nombre) in reconfig_norm
        con.execute("""
            INSERT INTO modelo_operacional_cruce
            (cruce_id,tipo_modelo,usa_reconfiguracion,version_prog_id,descripcion,fuente)
            VALUES (?,?,?,?,?,?)
        """, (
            cruce_id,
            'RECONFIG' if usa else 'NOREPROG',
            1 if usa else 0,
            2 if usa else 1,
            DESC_RECONFIG if usa else DESC_NOREPROG,
            'Regla de proyecto: 8 cruces con reconfiguracion; resto noreprog/base',
        ))

def crear_db(ruta: Path, schema: Path) -> sqlite3.Connection:
    """Crea una base nueva aplicando el esquema SQL indicado."""
    if ruta.exists():
        ruta.unlink()
    con = sqlite3.connect(ruta)
    con.executescript(schema.read_text(encoding='utf-8'))
    return con


# --------------------------------------------------------------------- #
#  carga de cada base
# --------------------------------------------------------------------- #
def cargar_infraestructura(wb_base, wb_reprog) -> dict:
    """Construye infraestructura.db. Devuelve mapas auxiliares."""
    con = crear_db(DIR_DATA / 'infraestructura.db',
                   DIR_SCHEMA / '1_infraestructura.sql')
    con.execute('PRAGMA foreign_keys = ON')

    # --- estaciones (orden de la linea, desde la hoja Itinerario) ---
    it = wb_base['Itinerario (INPUT)']
    estaciones, vistas = [], set()
    for r in range(12, it.max_row + 1):
        nom = it.cell(row=r, column=1).value
        if nom and norm(nom) not in vistas:
            vistas.add(norm(nom))
            estaciones.append(str(nom).strip())

    bbdd = wb_base['BBDD']
    # estaciones referenciadas en BBDD que no esten en el itinerario
    for r in range(2, bbdd.max_row + 1):
        if str(bbdd.cell(row=r, column=2).value or '') == 'Observaciones':
            break
        for col in (11, 12, 13, 14, 15, 16):  # K,L,M,N,O,P
            nom = bbdd.cell(row=r, column=col).value
            if nom and norm(nom) not in vistas:
                vistas.add(norm(nom))
                estaciones.append(str(nom).strip())

    est_id = {}
    for i, nom in enumerate(estaciones, start=1):
        con.execute('INSERT INTO estaciones (estacion_id,nombre,orden_linea) '
                    'VALUES (?,?,?)', (i, nom, i))
        est_id[norm(nom)] = i

    # --- cruces + parametros_barrera + cruce_tramo (hoja BBDD) ---
    cruce_id = {}
    alarma = to_seconds(bbdd.cell(row=25, column=7).value)  # G25
    for r in range(2, bbdd.max_row + 1):
        cid = bbdd.cell(row=r, column=1).value
        nombre = bbdd.cell(row=r, column=2).value
        if cid is None or str(nombre or '') == 'Observaciones':
            break
        sent = bbdd.cell(row=r, column=21).value          # U
        sent = sent if sent in ('CC', 'CW') else None
        con.execute(
            'INSERT INTO cruces (cruce_id,nombre,comuna,latitud,longitud,'
            'num_pistas_total,num_carriles_lateral,tiene_semaforo,'
            'afecta_lateral,sentido_afectacion,estacion_cercana_id,'
            'dist_estacion_m,estado_camaras) VALUES (?,?,?,?,?,?,?,?,?,?,?,?,?)',
            (int(cid), str(nombre).strip(),
             bbdd.cell(row=r, column=3).value,
             to_float_cl(bbdd.cell(row=r, column=4).value),
             to_float_cl(bbdd.cell(row=r, column=5).value),
             bbdd.cell(row=r, column=6).value,
             2,
             1 if str(bbdd.cell(row=r, column=7).value).lower() == 'si' else 0,
             1 if str(bbdd.cell(row=r, column=20).value).lower() == 'si' else 0,
             sent,
             est_id.get(norm(bbdd.cell(row=r, column=11).value)),
             to_float_cl(bbdd.cell(row=r, column=9).value),
             bbdd.cell(row=r, column=8).value))
        cruce_id[norm(nombre)] = int(cid)

        # tiempos de barrera: V = CW-CC, W = CC-CW
        for sentido, col in (('CW', 22), ('CC', 23)):
            t = to_seconds(bbdd.cell(row=r, column=col).value)
            if t is not None:
                con.execute(
                    'INSERT INTO parametros_barrera (cruce_id,sentido,'
                    'tiempo_barrera_s,margen_pre_s,margen_post_s,'
                    'tiempo_alarma_s,fuente) VALUES (?,?,?,?,?,?,?)',
                    (int(cid), sentido, t, 10, 10, alarma, 'BBDD xlsx'))
        # tramo: CW -> M,N ; CC -> O,P
        dist_tot = to_float_cl(bbdd.cell(row=r, column=10).value)
        dist_des = to_float_cl(bbdd.cell(row=r, column=9).value)
        for sentido, c1, c2 in (('CW', 13, 14), ('CC', 15, 16)):
            con.execute(
                'INSERT INTO cruce_tramo (cruce_id,sentido,estacion_desde_id,'
                'estacion_hasta_id,dist_desde_m,dist_total_m) VALUES (?,?,?,?,?,?)',
                (int(cid), sentido,
                 est_id.get(norm(bbdd.cell(row=r, column=c1).value)),
                 est_id.get(norm(bbdd.cell(row=r, column=c2).value)),
                 dist_des, dist_tot))

    # --- versiones de programacion + planes + fases ---
    versiones = [(1, 'Base real',         wb_base),
                 (2, 'Reprogramado Mar9', wb_reprog)]
    for vid, vnom, wb in versiones:
        con.execute('INSERT INTO versiones_programacion '
                    '(version_prog_id,nombre) VALUES (?,?)', (vid, vnom))
        pl = wb['Plan']
        for r in range(2, pl.max_row + 1):
            ini = to_seconds(pl.cell(row=r, column=2).value)
            fin = to_seconds(pl.cell(row=r, column=3).value)
            pid = pl.cell(row=r, column=4).value
            if ini is None or pid is None:
                continue
            con.execute('INSERT INTO planes_horarios (version_prog_id,'
                        'hora_inicio_s,hora_fin_s,plan_id) VALUES (?,?,?,?)',
                        (vid, ini, min(fin, 86400), int(pid)))
        pf = wb['PROG_FASES']
        for r in range(3, pf.max_row + 1):
            cross = pf.cell(row=r, column=1).value
            if cross is None:
                continue
            cid = cruce_id.get(norm(cross))
            if cid is None:
                continue
            con.execute(
                'INSERT INTO programacion_fases (version_prog_id,cruce_id,'
                'plan_id,fase_id,duracion_s,entreverde_s,cum_inicio_s,'
                'cum_fin_s,es_verde_lateral,ciclo_s) '
                'VALUES (?,?,?,?,?,?,?,?,?,?)',
                (vid, cid,
                 int(pf.cell(row=r, column=2).value),
                 int(pf.cell(row=r, column=3).value),
                 int(pf.cell(row=r, column=4).value),
                 int(pf.cell(row=r, column=5).value or 0),
                 int(pf.cell(row=r, column=7).value),   # cum_inicio_s = cumstart (col G)
                 int(pf.cell(row=r, column=6).value),   # cum_fin_s    = cumend  (col F)
                 int(pf.cell(row=r, column=8).value or 0),
                 int(pf.cell(row=r, column=9).value)))
    poblar_modelo_operacional(con)
    con.commit()
    n_est, n_cru = (con.execute(f'SELECT count(*) FROM {t}').fetchone()[0]
                    for t in ('estaciones', 'cruces'))
    con.close()
    print(f'  infraestructura.db : {n_est} estaciones, {n_cru} cruces')
    return {'cruce_id': cruce_id, 'est_id': est_id}


def cargar_demanda(wb_base, wb_reprog, cruce_id) -> None:
    """Construye demanda.db."""
    con = crear_db(DIR_DATA / 'demanda.db', DIR_SCHEMA / '2_demanda.sql')
    con.execute('PRAGMA foreign_keys = ON')

    # --- campanias de aforo + llegadas (flujo CRUDO, sin k_dem) ---
    campanias = [(1, 'Aforos base (NOREPROG)',  wb_base),
                 (2, 'Aforos reprog (Mar9)',    wb_reprog)]
    n_lleg = 0
    for cid, cnom, wb in campanias:
        con.execute('INSERT INTO campanias_medicion (campania_id,nombre,'
                    'descripcion) VALUES (?,?,?)',
                    (cid, cnom, 'Importado del modelo Excel'))
        lg = wb['Llegadas']
        for r in range(2, lg.max_row + 1):
            cruce = lg.cell(row=r, column=1).value
            if cruce is None:
                continue
            kid = cruce_id.get(norm(cruce))
            t_ini = to_seconds(lg.cell(row=r, column=2).value)
            t_fin = to_seconds(lg.cell(row=r, column=3).value)
            veh_h = lg.cell(row=r, column=4).value
            if kid is None or t_ini is None or veh_h is None:
                continue
            con.execute(
                'INSERT OR IGNORE INTO llegadas_vehiculares (campania_id,'
                'cruce_id,t_inicio_s,t_fin_s,flujo_veh_h) VALUES (?,?,?,?,?)',
                (cid, kid, t_ini, min(t_fin, 86400), float(veh_h)))
            n_lleg += 1

    # --- itinerario (version unica) + eventos de barrera (HCALL) ---
    con.execute('INSERT INTO itinerario_versiones (itinerario_id,nombre,'
                'descripcion) VALUES (1,?,?)',
                ('Itinerario L2 base', 'Malla operacional del modelo Excel'))
    hc = wb_base['HCALL']
    n_ev = 0
    for nombre, (rin, rout) in HCALL_ROWS.items():
        kid = cruce_id.get(norm(nombre))
        if kid is None:
            continue
        ins = sorted(s for s in (to_seconds(hc.cell(row=rin, column=c).value)
                                 for c in range(2, 121)) if s is not None)
        outs = sorted(s for s in (to_seconds(hc.cell(row=rout, column=c).value)
                                  for c in range(2, 121)) if s is not None)
        for i in range(min(len(ins), len(outs))):
            con.execute(
                'INSERT INTO eventos_barrera (itinerario_id,cruce_id,'
                'instante_paso_s,hcall_in_s,hcall_out_s) VALUES (1,?,?,?,?)',
                (kid, outs[i] - 10, ins[i], outs[i]))
            n_ev += 1
    con.commit()
    con.close()
    print(f'  demanda.db         : {n_lleg} bandas de aforo, {n_ev} eventos HCALL')


def cargar_escenarios(cruce_id) -> None:
    """Construye escenarios.db con corridas de ejemplo (sin resultados)."""
    con = crear_db(DIR_DATA / 'escenarios.db', DIR_SCHEMA / '3_escenarios.sql')
    ejemplos = ['Costa Verde', 'Diagonal Bio Bio',
                'Portal San Pedro', 'Conavicop']
    eid = 0
    for nombre in ejemplos:
        kid = cruce_id.get(norm(nombre))
        if kid is None:
            continue
        eid += 1
        con.execute(
            'INSERT INTO escenarios (escenario_id,nombre,cruce_id,'
            'version_prog_id,campania_id,itinerario_id,k_dem,modo,notas) '
            'VALUES (?,?,?,?,?,?,?,?,?)',
            (eid, f'{nombre} - base', kid, 1, 1, 1, 1.1, 'corrected',
             'Escenario de ejemplo generado por la migracion'))
    con.commit()
    con.close()
    print(f'  escenarios.db      : {eid} escenarios de ejemplo')


def main():
    f_base, f_reprog = DIR_FUENTES / XLSX_BASE, DIR_FUENTES / XLSX_REPROG
    for f in (f_base, f_reprog):
        if not f.exists():
            raise SystemExit(
                f'ERROR: falta {f.name} en la carpeta fuentes/.\n'
                'Copie ahi los dos .xlsx originales y vuelva a ejecutar.')
    print('Leyendo libros Excel (puede tardar ~1 min)...')
    wb_base   = openpyxl.load_workbook(f_base,   data_only=True)
    wb_reprog = openpyxl.load_workbook(f_reprog, data_only=True)
    print('Construyendo bases de datos:')
    aux = cargar_infraestructura(wb_base, wb_reprog)
    cargar_demanda(wb_base, wb_reprog, aux['cruce_id'])
    cargar_escenarios(aux['cruce_id'])
    wb_base.close()
    wb_reprog.close()
    print('Listo. Bases generadas en data/.')


if __name__ == '__main__':
    main()
